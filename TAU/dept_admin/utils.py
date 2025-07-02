import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import transaction
from django.contrib.auth.models import User
from django.utils import timezone
import logging
from openpyxl.worksheet.datavalidation import DataValidation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.crypto import get_random_string
from django.conf import settings
from django.urls import reverse
from core.models import Profile

logger = logging.getLogger(__name__)

def create_excel_template():
    """Create an Excel template for bulk student creation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Details"

    # Define headers
    headers = ['Roll Number', 'First Name', 'Last Name', 'Phone Number']
    
    # Style headers
    header_font = Font(bold=True)
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        ws.column_dimensions[chr(64 + col)].width = 20

    # Add data validation for Roll Number (12 digits)
    dv = DataValidation(type="textLength", operator="equal", formula1=12)
    dv.error = "Roll Number must be exactly 12 digits"
    dv.errorTitle = "Invalid Roll Number"
    ws.add_data_validation(dv)
    dv.add(f"A2:A1000")  # Apply to Roll Number column

    # Create the file in memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def check_user_profile_status(roll_number):
    """Debug function to check the status of a user and their profile."""
    from core.models import Profile
    from django.contrib.auth.models import User
    
    try:
        user = User.objects.get(username=roll_number)
        print(f"User found: {user.username} ({user.email})")
        
        # Check if user has profile attribute
        has_profile_attr = hasattr(user, 'profile')
        print(f"Has profile attribute: {has_profile_attr}")
        
        # Check if profile exists in database
        try:
            profile = Profile.objects.get(user=user)
            print(f"Profile found in database: {profile.id}")
            print(f"Profile department: {profile.department.name if profile.department else 'None'}")
            print(f"Profile is_admin: {profile.is_admin}")
        except Profile.DoesNotExist:
            print("Profile not found in database")
        
        # Check if profile exists by user_id
        profile_by_id = Profile.objects.filter(user_id=user.id).first()
        if profile_by_id:
            print(f"Profile found by user_id: {profile_by_id.id}")
        else:
            print("Profile not found by user_id")
            
    except User.DoesNotExist:
        print(f"User not found: {roll_number}")

def process_excel_file(file, department):
    """Process the uploaded Excel file and create student accounts."""
    try:
        logger.info("Starting to process Excel file")
        
        # Read the Excel file
        df = pd.read_excel(file, dtype=str)  # Read all columns as strings
        logger.info(f"Read Excel file. Shape: {df.shape}")
        
        # Clean column names and data
        df.columns = df.columns.str.strip()
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].str.strip()
        
        # Verify required columns
        required_columns = ['Roll Number', 'First Name', 'Last Name', 'Phone Number']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Missing required columns: {', '.join(missing_columns)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        created_students = []
        errors = []
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Get and validate roll number
                roll_number = str(row['Roll Number'])
                if not roll_number or not roll_number.isdigit() or len(roll_number) != 12:
                    errors.append(f"Row {index + 2}: Invalid roll number format. Must be exactly 12 digits.")
                    continue

                # Get and validate names
                first_name = str(row['First Name'])
                last_name = str(row['Last Name'])
                if not first_name or not last_name:
                    errors.append(f"Row {index + 2}: First name and last name are required.")
                    continue

                # Get and validate phone number
                phone_number = str(row['Phone Number']).strip()
                if not phone_number or not phone_number.isdigit() or len(phone_number) != 10:
                    errors.append(f"Row {index + 2}: Invalid phone number. Must be a 10-digit number.")
                    continue

                # Generate email
                email = f"{roll_number}@apollouniversity.edu.in"
                
                # Create or update user and profile
                with transaction.atomic():
                    # Check if user exists
                    user, created = User.objects.get_or_create(
                        username=roll_number,
                        defaults={
                            'email': email,
                            'first_name': first_name,
                            'last_name': last_name,
                            'is_active': True
                        }
                    )

                    if not created:
                        # Update existing user
                        user.email = email
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_active = True
                    # Set/reset password
                    user.set_password('Random@123')
                    user.save()
                    # Create or update profile
                    profile, _ = Profile.objects.get_or_create(
                        user=user,
                        defaults={
                            'department': department,
                            'is_admin': False,
                            'must_change_password': True,
                            'phone_number': phone_number
                        }
                    )
                    # Update profile if it existed
                    profile.department = department
                    profile.is_admin = False
                    profile.must_change_password = True
                    profile.phone_number = phone_number
                    profile.save()
                    from .utils import send_registration_sms
                    send_registration_sms(phone_number, 'https://apollouniversity.edu.in/login')
                    created_students.append({
                        'roll_number': roll_number,
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email
                    })
                    logger.info(f"Successfully processed student: {roll_number}")
            except Exception as e:
                error_msg = f"Row {index + 2}: Error processing student: {str(e)}"
                logger.error(error_msg)
                errors.append(error_msg)
                continue
        return created_students, errors
        
    except Exception as e:
        error_msg = f"Error processing Excel file: {str(e)}"
        logger.error(error_msg)
        raise ValueError(error_msg)

def send_registration_email(email, roll_number, registration_link):
    """Send registration email to student."""
    subject = 'Complete Your TAU Registration'
    html_message = render_to_string('dept_admin/emails/registration_email.html', {
        'roll_number': roll_number,
        'registration_link': registration_link
    })
    
    try:
        send_mail(
            subject=subject,
            message=strip_tags(html_message),
            html_message=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )
        logger.info(f"Registration email sent to {email}")
        return True
    except Exception as e:
        logger.error(f"Failed to send registration email to {email}: {str(e)}")
        return False

def process_student_registrations(file, department=None):
    """Process student registrations from Excel file and send registration emails."""
    try:
        logger.info("Starting to process student registrations")
        
        # Read the Excel file
        df = pd.read_excel(file, dtype=str, header=0)
        logger.info(f"Read Excel file. Shape: {df.shape}")
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Verify required columns
        required_columns = ['Roll Number', 'First Name', 'Last Name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        processed_students = []
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get and clean values
                roll_number = str(row['Roll Number']).strip()
                first_name = str(row['First Name']).strip()
                last_name = str(row['Last Name']).strip()
                email = f"{roll_number}@apollouniversity.edu.in"
                
                # Validate data
                if not roll_number or not roll_number.isdigit() or len(roll_number) != 12:
                    errors.append(f"Row {index + 2}: Invalid roll number format")
                    continue
                
                if not first_name or not last_name:
                    errors.append(f"Row {index + 2}: First name and last name are required")
                    continue
                
                # Generate registration token
                registration_token = get_random_string(64)
                
                with transaction.atomic():
                    # Create or update temporary registration record
                    from .models import PendingRegistration
                    registration, created = PendingRegistration.objects.update_or_create(
                        roll_number=roll_number,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'token': registration_token,
                            'department': department,
                            'expires_at': timezone.now() + timezone.timedelta(days=7)
                        }
                    )
                    
                    # Generate registration link
                    registration_link = f"{settings.SITE_URL}{reverse('student:complete_registration')}?token={registration_token}"
                    
                    # Send registration email
                    if send_registration_email(email, roll_number, registration_link):
                        processed_students.append({
                            'roll_number': roll_number,
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'status': 'Email sent'
                        })
                    else:
                        errors.append(f"Row {index + 2}: Failed to send registration email")
            
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue
        
        return processed_students, errors
    
    except Exception as e:
        raise ValueError(f"Error processing file: {str(e)}")

def complete_student_registration(token, password):
    """Complete student registration with the provided token and password."""
    try:
        with transaction.atomic():
            # Get pending registration
            pending_reg = PendingRegistration.objects.get(
                token=token,
                expires_at__gt=timezone.now(),
                is_used=False
            )
            
            # Create user account
            user = User.objects.create_user(
                username=pending_reg.roll_number,
                email=pending_reg.email,
                password=password,
                first_name=pending_reg.first_name,
                last_name=pending_reg.last_name,
                is_active=True
            )
            
            # Create profile
            Profile.objects.create(
                user=user,
                department=pending_reg.department,
                is_admin=False,
                must_change_password=False  # Not needed since they set their own password
            )
            
            # Mark registration as used
            pending_reg.is_used = True
            pending_reg.save()
            
            return True, user
            
    except PendingRegistration.DoesNotExist:
        return False, "Invalid or expired registration link"
    except Exception as e:
        logger.error(f"Error completing registration: {str(e)}")
        return False, "An error occurred during registration"

def send_registration_sms(phone_number, login_url):
    """Send registration SMS to student (placeholder)."""
    message = f"Your Apollo University account has been created. You may now log in at: {login_url}"
    # TODO: Integrate with real SMS gateway here
    print(f"SMS to {phone_number}: {message}")
    logger.info(f"SMS to {phone_number}: {message}")

# ... (code after process_excel_file, if any) 